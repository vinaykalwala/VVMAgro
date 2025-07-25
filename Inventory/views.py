import json
from django.shortcuts import render
from .models import *
from .forms import *

# Create your views here.
def home(request):
    return render(request,'staticpages/home.html')

def aboutus(request):
    return render(request,'staticpages/aboutus.html')    

def products(request):
    return render(request,'staticpages/products.html')

def front_dozer_tractor(request):
    return render(request, 'staticpages/front_dozer_tractor.html')

def hydraulicboomlift(request):
    return render(request, 'staticpages/hydraulicboomlift.html')

def hole_digger(request):
    return render(request, 'staticpages/hole_digger.html')
    
def frontendloader(request):
    return render(request, 'staticpages/frontendloader.html')

def lift_tractor(request):
    return render(request, 'staticpages/lift_tractor.html')



def branchesanddistributors(request):
    return render(request,'staticpages/branchesanddistributors.html')

def careers(request):
    return render(request,'staticpages/careers.html')

def tractorproducts(request):
    return render(request,'staticpages/farmingtractor.html')  

def pesticideprayer(request):
    return render(request,'staticpages/pesticidesprayer.html')    

    


from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from .forms import CustomUserCreationForm, CustomAuthenticationForm

@login_required
def signup_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('signup')
    else:
        form = CustomUserCreationForm()
    return render(request, 'backendpages/signup.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect based on role
                if user.is_superuser:
                    return redirect('admin_dashboard')
                elif user.role == 'user':
                    return redirect('user_dashboard')
                else:
                    return redirect('manager_dashboard')
        else:
            # Invalid login
            return render(request, 'backendpages/login.html', {'form': form, 'error': 'Invalid credentials'})
    else:
        form = CustomAuthenticationForm()
    return render(request, 'backendpages/login.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def admin_dashboard(request):
    return render(request, 'backendpages/admin_dashboard.html',{'user_profile': request.user})

@login_required
def user_dashboard(request):
    return render(request, 'backendpages/user_dashboard.html',{'user_profile': request.user})

@login_required
def manager_dashboard(request):
    return render(request, 'backendpages/manager_dashboard.html',{'user_profile': request.user})


from django.shortcuts import render, get_object_or_404
from .models import Voucher
from num2words import num2words

from django.db.models import Q
from datetime import datetime

@login_required
def voucher_list_view(request):
    vouchers = Voucher.objects.all().order_by('-created_at')

    voucher_number = request.GET.get('voucher_number', '').strip()
    voucher_type = request.GET.get('voucher_type', '').strip()
    party_name = request.GET.get('party_name', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    date_exact = request.GET.get('date_exact', '').strip()

    if voucher_number:
        vouchers = vouchers.filter(voucher_number__icontains=voucher_number)

    if voucher_type:
        vouchers = vouchers.filter(voucher_type=voucher_type)

    if party_name:
        vouchers = vouchers.filter(party__name__icontains=party_name)

    if date_exact:
        try:
            date_exact_obj = datetime.strptime(date_exact, '%Y-%m-%d')
            vouchers = vouchers.filter(created_at__date=date_exact_obj)
        except ValueError:
            pass
    else:
        # Only apply range filters if exact date not specified
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                vouchers = vouchers.filter(created_at__date__gte=date_from_obj)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                vouchers = vouchers.filter(created_at__date__lte=date_to_obj)
            except ValueError:
                pass

    context = {
        'vouchers': vouchers,
        'search_params': {
            'voucher_number': voucher_number,
            'voucher_type': voucher_type,
            'party_name': party_name,
            'date_from': date_from,
            'date_to': date_to,
            'date_exact': date_exact,
        }
    }
    return render(request, 'voucher_list.html', context)

@login_required
def voucher_detail_view(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)
    items = voucher.items.all()

    # Determine label based on movement_type of first item
    if voucher.voucher_type == 'Delivery_Challan':
        party_label = 'Delivery Challan'
    elif voucher.voucher_type == 'Quotation':
        party_label = 'Quotation'
    else:
        if items.exists():
            movement_type = voucher.movement_type
            if movement_type == 'in':
                party_label = 'Seller'
            elif movement_type == 'out':
                party_label = 'Buyer'
            elif movement_type == 'job_work':
                party_label = 'Vendor'
            else:
                party_label = 'Party'
        else:
            party_label = 'Party'

    total_subtotal = sum(item.subtotal for item in voucher.items.all())
    total_cgst = sum(item.cgst_amount for item in voucher.items.all())
    total_sgst = sum(item.sgst_amount for item in voucher.items.all())
    total_igst = sum(item.igst_amount for item in voucher.items.all())
    grand_total = sum(item.total_amount for item in voucher.items.all())
    amount_in_words = f"Rupees {num2words(voucher.grand_total, lang='en_IN').title()} "

    context = {
        'voucher': voucher,
        'items': items,
        'party_label': party_label,
        'total_subtotal': total_subtotal,
        'total_cgst': total_cgst,
        'total_sgst': total_sgst,
        'total_igst': total_igst,
        'grand_total': grand_total,
        'amount_in_words': amount_in_words,
    }
    return render(request, 'voucher_detail.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from .models import ProductGroup, Warehouse, Product, PartyGroup, Party
from .forms import ProductGroupForm, WarehouseForm, ProductForm, PartyGroupForm, PartyForm

@login_required
# ProductGroup Views
def productgroup_list(request):
    groups = ProductGroup.objects.all()
    return render(request, 'productgroup_list.html', {'groups': groups})

@login_required
def productgroup_create(request):
    form = ProductGroupForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('productgroup_list')
    return render(request, 'productgroup_form.html', {'form': form})

@login_required
def productgroup_update(request, pk):
    group = get_object_or_404(ProductGroup, pk=pk)
    form = ProductGroupForm(request.POST or None, instance=group)
    if form.is_valid():
        form.save()
        return redirect('productgroup_list')
    return render(request, 'productgroup_form.html', {'form': form})

@login_required
def productgroup_delete(request, pk):
    group = get_object_or_404(ProductGroup, pk=pk)
    if request.method == 'POST':
        group.delete()
        return redirect('productgroup_list')
    return render(request, 'confirm_delete.html', {'object': group})

@login_required
# Warehouse Views
def warehouse_list(request):
    warehouses = Warehouse.objects.all()
    return render(request, 'warehouse_list.html', {'warehouses': warehouses})

@login_required
def warehouse_create(request):
    form = WarehouseForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('warehouse_list')
    return render(request, 'warehouse_form.html', {'form': form})

@login_required
def warehouse_update(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    form = WarehouseForm(request.POST or None, instance=warehouse)
    if form.is_valid():
        form.save()
        return redirect('warehouse_list')
    return render(request, 'warehouse_form.html', {'form': form})

@login_required
def warehouse_delete(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        warehouse.delete()
        return redirect('warehouse_list')
    return render(request, 'confirm_delete.html', {'object': warehouse})

# Product Views
from django.db.models import Q
@login_required
def product_list(request):
    query = request.GET.get('q', '').strip()
    products = Product.objects.select_related('group', 'warehouse').all()

    if query:
        products = products.filter(
            Q(name__icontains=query) |
            Q(group__name__icontains=query) |
            Q(warehouse__name__icontains=query) |
            Q(stock_quantity__icontains=query) |
            Q(hsn_sac__icontains=query)
        )

    return render(request, 'product_list.html', {
        'products': products,
        'query': query,
    })

@login_required
def product_create(request):
    form = ProductForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('product_list')
    return render(request, 'product_form.html', {'form': form})

@login_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = ProductForm(request.POST or None, instance=product)
    if form.is_valid():
        form.save()
        return redirect('product_list')
    return render(request, 'product_form.html', {'form': form})

@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'confirm_delete.html', {'object': product})

@login_required
# PartyGroup Views
def partygroup_list(request):
    groups = PartyGroup.objects.all()
    return render(request, 'partygroup_list.html', {'groups': groups})

@login_required
def partygroup_create(request):
    form = PartyGroupForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('partygroup_list')
    return render(request, 'partygroup_form.html', {'form': form})

@login_required
def partygroup_update(request, pk):
    group = get_object_or_404(PartyGroup, pk=pk)
    form = PartyGroupForm(request.POST or None, instance=group)
    if form.is_valid():
        form.save()
        return redirect('partygroup_list')
    return render(request, 'partygroup_form.html', {'form': form})

@login_required
def partygroup_delete(request, pk):
    group = get_object_or_404(PartyGroup, pk=pk)
    if request.method == 'POST':
        group.delete()
        return redirect('partygroup_list')
    return render(request, 'confirm_delete.html', {'object': group})

# Party Views
# views.py
from django.db.models import Q
from django.shortcuts import render
from .models import Party

@login_required
def party_list(request):
    query = request.GET.get('q', '').strip()
    parties = Party.objects.select_related('group').all()

    if query:
        parties = parties.filter(
            Q(name__icontains=query) |
            Q(group__name__icontains=query) |
            Q(gstin_uin_number__icontains=query) |
            Q(address__icontains=query) |
            Q(location__icontains=query) |
            Q(pincode__icontains=query) |
            Q(state__icontains=query) |
            Q(phone__icontains=query) |
            Q(email__icontains=query)
        )

    return render(request, 'party_list.html', {
        'parties': parties,
        'query': query
    })

@login_required
def party_create(request):
    form = PartyForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('party_list')
    return render(request, 'party_form.html', {'form': form})

@login_required
def party_update(request, pk):
    party = get_object_or_404(Party, pk=pk)
    form = PartyForm(request.POST or None, instance=party)
    if form.is_valid():
        form.save()
        return redirect('party_list')
    return render(request, 'party_form.html', {'form': form})

@login_required
def party_delete(request, pk):
    party = get_object_or_404(Party, pk=pk)
    if request.method == 'POST':
        party.delete()
        return redirect('party_list')
    return render(request, 'confirm_delete.html', {'object': party})



from .models import Job, JobApplication
from .forms import JobApplicationForm

def apply_for_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    
    if request.method == "POST":
        form = JobApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.job = job
            application.save()
            messages.success(request, "Your application has been submitted successfully!")
            return redirect("apply_for_job", job_id=job.id)  
        else:
            messages.error(request, "There was an error with your application. Please check the form.")
    else:
        form = JobApplicationForm()

    return render(request, "staticpages/apply.html", {"form": form, "job": job})

def careers(request):
    jobs = Job.objects.all().order_by('-created_at')
    return render(request, "staticpages/careers.html", {"jobs": jobs})

def contact_create_view(request):
    success = False  # Flag to show success message

    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            success = True
            form = ContactForm()  # Clear form after submission
    else:
        form = ContactForm()

    return render(request, 'contact_form.html', {'form': form, 'success': success})

@login_required
def contact_list_view(request):
    contacts = Contact.objects.all().order_by('-created_at')
    return render(request, 'contact_list.html', {'contacts': contacts})

from django.contrib.auth.decorators import login_required
from .forms import JobForm
from .models import JobApplication

@login_required
def job_post_view(request):
    if not (request.user.role == 'manager' or request.user.role == 'user' or request.user.is_superuser):
        messages.error(request, "Only admins can post jobs.")
        return redirect('home')
    
    if request.method == 'POST':
        form = JobForm(request.POST)
        if form.is_valid():
            job = form.save()
            messages.success(request, f"Job '{job.title}' posted successfully!")
            return redirect('job_applications')
    else:
        form = JobForm()
    
    return render(request, 'backendpages/job_post.html', {'form': form})

@login_required
def job_applications_view(request):
    if not (request.user.role == 'manager' or request.user.role == 'user' or request.user.is_superuser):
        messages.error(request, "Only admins can view applications.")
        return redirect('user_dashboard')
    
    applications = JobApplication.objects.all().order_by('-applied_at')
    return render(request, 'backendpages/job_applications.html', {'applications': applications})

@login_required
def application_detail_view(request, application_id):
    if not (request.user.role == 'manager' or request.user.role == 'user' or request.user.is_superuser):
        messages.error(request, "Only admins can view application details.")
        return redirect('user_dashboard')
    
    application = get_object_or_404(JobApplication, id=application_id)
    return render(request, 'backendpages/application_detail.html', {'application': application})


from django.db.models import Q

@login_required
def stock_report_view(request):
    query = request.GET.get('q', '')
    warehouse_id = request.GET.get('warehouse', '')
    phase = request.GET.get('phase', '')
    supply_type = request.GET.get('type_of_supply', '')

    products = Product.objects.select_related('warehouse').all()

    if query:
        products = products.filter(name__icontains=query)

    if warehouse_id:
        products = products.filter(warehouse_id=warehouse_id)

    if phase:
        products = products.filter(phase=phase)

    if supply_type:
        products = products.filter(type_of_supply=supply_type)

    stock_data = []

    for product in products:
        stock_data.append({
            "product_id": product.id,
            "product_name": product.name,
            "type_of_supply": product.get_type_of_supply_display(),
            "stock_quantity": product.stock_quantity,
            "warehouse_name": product.warehouse.name if product.warehouse else "No Warehouse Assigned",
            "warehouse_location": product.warehouse.location if product.warehouse and hasattr(product.warehouse, 'location') else "Location Not Available",
            "phase": product.get_phase_display() if product.type_of_supply == 'goods' else "N/A"
        })

    warehouses = Warehouse.objects.all()

    return render(request, 'stock_report.html', {
        'stock_data': stock_data,
        'warehouses': warehouses,
        'query': query,
        'selected_warehouse': warehouse_id,
        'selected_phase': phase,
        'selected_type': supply_type,
    })


from django.shortcuts import render, get_object_or_404
from .models import Voucher
from datetime import datetime, timedelta
from django.db.models import Q
from django.utils.timezone import make_aware


@login_required
def day_book_view(request):
    vouchers = Voucher.objects.all().order_by('-created_at')

    # Get filter parameters from GET request
    date = request.GET.get('date')
    week = request.GET.get('week')  # Format: YYYY-Www (e.g., 2025-W21)
    month = request.GET.get('month')  # Format: YYYY-MM
    year = request.GET.get('year')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    voucher_type = request.GET.get('voucher_type')

    # Apply filters
    if date:
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        vouchers = vouchers.filter(created_at__date=date_obj)

    if week:
        year_str, week_str = week.split('-W')
        year, week = int(year_str), int(week_str)
        start_of_week = datetime.strptime(f'{year}-W{week}-1', "%Y-W%W-%w")
        end_of_week = start_of_week + timedelta(days=6)
        vouchers = vouchers.filter(created_at__date__range=[start_of_week.date(), end_of_week.date()])

    if month:
        year, month = map(int, month.split('-'))
        start_of_month = datetime(year, month, 1)
        if month == 12:
            end_of_month = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_of_month = datetime(year, month + 1, 1) - timedelta(days=1)
        vouchers = vouchers.filter(created_at__date__range=[start_of_month.date(), end_of_month.date()])

    if year:
        vouchers = vouchers.filter(created_at__year=year)

    if start_date and end_date:
        start = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        end = make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1)
        vouchers = vouchers.filter(created_at__range=[start, end])

    if voucher_type:
        vouchers = vouchers.filter(voucher_type=voucher_type)

    context = {
        'vouchers': vouchers,
    }
    return render(request, 'day_book.html', context)


from django.shortcuts import render
from .models import Gallery

def gallery(request):
    images = Gallery.objects.all()
    return render(request, 'staticpages/gallery.html', {'images': images})

from django.shortcuts import render, redirect, get_object_or_404
from .models import Gallery
from .forms import GalleryForm
from django.contrib import messages

def gallery_list(request):
    galleries = Gallery.objects.all()
    return render(request, 'staticpages/gallery_list.html', {'galleries': galleries})

@login_required
def gallery_create(request):
    if request.method == 'POST':
        form = GalleryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gallery item created successfully!')
            return redirect('gallery_list')
    else:
        form = GalleryForm()
    return render(request, 'staticpages/gallery_form.html', {'form': form})

@login_required
def gallery_edit(request, pk):
    gallery = get_object_or_404(Gallery, pk=pk)
    if request.method == 'POST':
        form = GalleryForm(request.POST, request.FILES, instance=gallery)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gallery item updated successfully!')
            return redirect('gallery_list')
    else:
        form = GalleryForm(instance=gallery)
    return render(request, 'staticpages/gallery_form.html', {'form': form})

@login_required
def gallery_delete(request, pk):
    gallery = get_object_or_404(Gallery, pk=pk)
    if request.method == 'POST':
        gallery.delete()
        messages.success(request, 'Gallery item deleted successfully!')
        return redirect('gallery_list')
    return render(request, 'staticpages/gallery_confirm_delete.html', {'gallery': gallery})


from .forms import CustomUserEditForm, ChangePasswordForm
from django.contrib.auth.decorators import user_passes_test

def is_admin(user):
    return user.is_superuser or user.role == 'manager'


@user_passes_test(is_admin)
def user_list(request):
    users = CustomUser.objects.filter(is_superuser=False)
    return render(request, 'user_list.html', {'users': users})


@user_passes_test(is_admin)
def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        form = CustomUserEditForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully.')
            return redirect('user_list')
    else:
        form = CustomUserEditForm(instance=user)
    return render(request, 'edit_user.html', {'form': form, 'user_obj': user})


@user_passes_test(is_admin)
def change_user_password(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        form = ChangePasswordForm(request.POST)
        if form.is_valid():
            password = form.cleaned_data['new_password']
            user.set_password(password)
            user.original_password = password
            user.save()
            messages.success(request, 'Password changed successfully.')
            return redirect('user_list')
    else:
        form = ChangePasswordForm()
    return render(request, 'change_password.html', {'form': form, 'user_obj': user})


from django.shortcuts import render
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth import get_user_model
from .models import Product
from django.db.models import F

@login_required
def notify_low_stock_products(request):
    # Get products where stock_quantity < stock_limit
    low_stock_products = Product.objects.select_related('warehouse', 'group').filter(stock_quantity__lt=F('stock_limit'))

    # Compose and send email if any low-stock products
    if low_stock_products.exists():
        product_lines = []
        for product in low_stock_products:
            warehouse_name = product.warehouse.name if product.warehouse else "N/A"
            group_name = product.group.name if product.group else "N/A"
            line = (
                f"- {product.name} | Stock: {product.stock_quantity} | "
                f"Stock Limit: {product.stock_limit} | "
                f"Warehouse: {warehouse_name} | Group: {group_name} | Unit: {product.unit_of_measurement}"
            )
            product_lines.append(line)

        # Email content
        message_body = (
            "The following products have stock below their stock limit:\n\n" +
            "\n".join(product_lines)
        )

        # Get superusers with valid email
        User = get_user_model()
        superusers = User.objects.filter(is_superuser=True, email__isnull=False)

        # Send email to each superuser
        for superuser in superusers:
            send_mail(
                subject="Low Stock Alert",
                message=message_body,
                from_email=settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER,
                recipient_list=[superuser.email],
                fail_silently=False,
            )

    # Render the template with context (shows all low stock products in UI)
    return render(request, 'low_stock_products.html', {
        'low_stock_products': low_stock_products,
    })

@login_required
def voucher_type_list(request):
    types = [vt[0] for vt in Voucher.VOUCHER_TYPES]
    return render(request, 'voucher_type_list.html', {'voucher_types': types})

from django.db import transaction
from django.shortcuts import render, redirect
from django.forms import inlineformset_factory
from django.contrib import messages
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP
import json

from .models import Voucher, VoucherProductItem, Product
from .forms import VoucherForm, VoucherProductItemFormSet
from django.contrib.auth.decorators import login_required

@login_required
def create_voucher_with_items(request, voucher_type):
    movement_type_map = {
        'Seller_Voucher': 'in',
        'Buyer_Voucher': 'out',
        'Job_Work_Voucher': 'job_work',
        'Quotation': '',
        'Delivery_Challan': '',
    }

    voucher_type_prefix_map = {
        'Buyer_Voucher': 'INV',
        'Seller_Voucher': 'P',
        'Job_Work_Voucher': 'JW',
        'Quotation': 'Q',
        'Delivery_Challan': 'DC',
    }

    initial_movement_type = movement_type_map.get(voucher_type, '')

    # ✅ Always generate preview voucher number
    now = timezone.now()
    fy_start = now.year if now.month >= 4 else now.year - 1
    fy_end = fy_start + 1
    fy_str = f"{str(fy_start)[-2:]}-{str(fy_end)[-2:]}"
    prefix_code = voucher_type_prefix_map.get(voucher_type, 'UNK')
    prefix = f"VVM/{fy_str}/{prefix_code}/"
    count = Voucher.objects.filter(voucher_number__startswith=prefix).count() + 1
    preview_voucher_number = f"{prefix}{str(count).zfill(3)}"

    def round_to_nearest_10_with_00(value):
        value = Decimal(value)
        rounded = (value / 10).quantize(Decimal('1'), rounding=ROUND_HALF_UP) * 10
        return rounded.quantize(Decimal('1.00'))

    def save_voucher_and_items(form, formset, voucher_type):
        with transaction.atomic():
            voucher = form.save(commit=False)
            voucher.voucher_type = voucher_type
            voucher.movement_type = movement_type_map.get(voucher_type, '')

            if not voucher.voucher_number:
                voucher.voucher_number = f"{prefix}{str(count).zfill(3)}"

            voucher.save()

            total_subtotal = total_cgst = total_sgst = total_igst = grand_total = 0

            for item_form in formset:
                if item_form.cleaned_data.get('DELETE', False):
                    continue
                item = item_form.save(commit=False)
                item.voucher = voucher

                item.subtotal = item.quantity * item.price_per_unit
                item.cgst_amount = (item.subtotal * item.cgst_percent) / 100
                item.sgst_amount = (item.subtotal * item.sgst_percent) / 100
                item.igst_amount = (item.subtotal * item.igst_percent) / 100
                item.total_amount = item.subtotal + item.cgst_amount + item.sgst_amount + item.igst_amount

                if voucher_type == 'Seller_Voucher':
                    voucher.movement_type = 'in'
                    item.product.stock_quantity += int(item.quantity)
                    item.product.save()

                elif voucher_type == 'Buyer_Voucher':
                    voucher.movement_type = 'out'
                    item.product.stock_quantity -= int(item.quantity)
                    item.product.save()

                elif voucher_type == 'Job_Work_Voucher':
                    voucher.movement_type = 'job_work'
                    item.product.stock_quantity += int(item.quantity)
                    selected_phase = item_form.cleaned_data.get('phase', '')
                    if selected_phase:
                        item.product.phase = selected_phase
                        item.product.save()

                elif voucher_type in ['Quotation', 'Delivery_Challan']:
                    voucher.movement_type = ''

                item.save()

                total_subtotal += item.subtotal
                total_cgst += item.cgst_amount
                total_sgst += item.sgst_amount
                total_igst += item.igst_amount
                grand_total += item.total_amount

            if voucher.freight_applicable and voucher.freight_charge:
                grand_total += voucher.freight_charge

            # 💡 Round to nearest 10 and calculate round-off
            rounded_grand_total = round_to_nearest_10_with_00(grand_total)
            round_off = rounded_grand_total - grand_total

            # 🔄 Save all calculated values
            voucher.total_subtotal = total_subtotal
            voucher.total_cgst = total_cgst
            voucher.total_sgst = total_sgst
            voucher.total_igst = total_igst
            voucher.round_off_on_sales = round_off
            voucher.grand_total = rounded_grand_total
            voucher.save()

        return voucher

    if request.method == 'POST':
        form = VoucherForm(request.POST or None, voucher_type=voucher_type)
        formset = VoucherProductItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():

            if voucher_type == 'Job_Work_Voucher':
                invalid_products = []
                for item_form in formset:
                    if item_form.cleaned_data.get('DELETE', False):
                        continue
                    product = item_form.cleaned_data.get('product')
                    if product and product.type_of_supply != 'raw_material':
                        invalid_products.append(product.name)

                if invalid_products:
                    messages.error(request,
                        "For Job Work Voucher, only products with type_of_supply as 'raw_material' are allowed. "
                        "Invalid products: " + ", ".join(invalid_products)
                    )
                else:
                    voucher = save_voucher_and_items(form, formset, voucher_type)
                    return redirect('voucher_detail', voucher_id=voucher.id)

            elif voucher_type == 'Buyer_Voucher':
                insufficient_stock_products = []
                for item_form in formset:
                    if item_form.cleaned_data.get('DELETE', False):
                        continue
                    product = item_form.cleaned_data.get('product')
                    quantity = item_form.cleaned_data.get('quantity')
                    if product and quantity and product.stock_quantity < int(quantity):
                        insufficient_stock_products.append(
                            f"{product.name} (available: {product.stock_quantity}, requested: {int(quantity)})"
                        )

                if insufficient_stock_products:
                    messages.error(request,
                        "Insufficient stock for the following products: " + ", ".join(insufficient_stock_products)
                    )
                else:
                    voucher = save_voucher_and_items(form, formset, voucher_type)
                    return redirect('voucher_detail', voucher_id=voucher.id)

            else:
                voucher = save_voucher_and_items(form, formset, voucher_type)
                return redirect('voucher_detail', voucher_id=voucher.id)

    else:
        form = VoucherForm(initial={'voucher_type': voucher_type, 'movement_type': initial_movement_type})
        formset = VoucherProductItemFormSet()

    products = Product.objects.select_related('warehouse').all()
    product_data = {
        str(product.id): {
            "warehouse_id": str(product.warehouse.id) if product.warehouse else ""
        }
        for product in products
    }

    return render(request, 'voucher_form_with_items.html', {
        'form': form,
        'formset': formset,
        'voucher_type': voucher_type,
        "product_data": json.dumps(product_data),
        "preview_voucher_number": preview_voucher_number
    })


import pandas as pd
from .forms import ProductUploadForm
from .models import Product, ProductGroup, Warehouse


@login_required
def upload_products(request):
    if request.method == 'POST':
        form = ProductUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                df = pd.read_excel(excel_file)

                for _, row in df.iterrows():
                    group_obj = ProductGroup.objects.get(name=row['group'])
                    warehouse_obj = None
                    if 'warehouse' in row and pd.notna(row['warehouse']):
                        warehouse_obj = Warehouse.objects.filter(name=row['warehouse']).first()

                    # Handle stock_limit safely — convert to int, fallback to 0 if missing
                    stock_limit_value = row.get('stock_limit', 0)
                    if pd.isna(stock_limit_value):
                        stock_limit_value = 0
                    stock_limit_value = int(stock_limit_value)

                    # Handle stock_quantity safely — convert to int, fallback to 0 if missing
                    stock_quantity_value = row.get('stock_quantity', 0)
                    if pd.isna(stock_quantity_value):
                        stock_quantity_value = 0
                    stock_quantity_value = int(stock_quantity_value)

                    Product.objects.create(
                        name=row['name'],
                        group=group_obj,
                        price_per_unit=row['price_per_unit'],
                        unit_of_measurement=row['unit_of_measurement'],
                        stock_quantity=stock_quantity_value,  # Now properly handled as integer
                        warehouse=warehouse_obj,
                        stock_limit=stock_limit_value,

                        hsn_sac=row['hsn_sac'],
                        hsn_sac_details=row.get('hsn_sac_details', ''),
                        hsn_sac_source_of_details=row.get('hsn_sac_source_of_details', ''),
                        hsn_sac_description=row.get('hsn_sac_description', ''),
                        type_of_supply=row['type_of_supply'],
                        phase=row.get('phase', None),
                        description=row.get('description', '')
                    )

                messages.success(request, "Products uploaded successfully!")
                return redirect('upload_products')

            except Exception as e:
                messages.error(request, f"Error: {e}")

    else:
        form = ProductUploadForm()

    return render(request, 'upload_products.html', {'form': form})


from .forms import PartyUploadForm
@login_required
def upload_parties(request):
    if request.method == 'POST':
        form = PartyUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                df = pd.read_excel(excel_file)

                for _, row in df.iterrows():
                    try:
                        group = PartyGroup.objects.get(name=row['group'])
                    except PartyGroup.DoesNotExist:
                        messages.error(request, f"Group '{row['group']}' does not exist.")
                        continue

                    Party.objects.create(
                        group=group,
                        name=row['name'],
                        gstin_uin_number=row['gstin_uin_number'],
                        address=row['address'],
                        location=row['location'],
                        pincode=row['pincode'],
                        state=row['state'],
                        phone=row.get('phone', ''),
                        email=row.get('email', ''),
                        registration_type=row.get('registration_type', '')  # <- Added here

                    )

                messages.success(request, "Party data uploaded successfully!")
                return redirect('upload_parties')

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
    else:
        form = PartyUploadForm()
    
    return render(request, 'upload_parties.html', {'form': form})

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from decimal import Decimal, ROUND_HALF_UP
import json
from .models import Voucher, VoucherProductItem, Product
from .forms import VoucherForm, VoucherProductItemFormSet

@login_required
def edit_voucher_with_items(request, voucher_id):
    voucher = get_object_or_404(Voucher, id=voucher_id)
    voucher_type = voucher.voucher_type
    
    movement_type_map = {
        'Seller_Voucher': 'in',
        'Buyer_Voucher': 'out',
        'Job_Work_Voucher': 'job_work',
        'Quotation': '',
        'Delivery_Challan': '',
    }

    def round_to_nearest_10_with_00(value):
        value = Decimal(value)
        rounded = (value / 10).quantize(Decimal('1'), rounding=ROUND_HALF_UP) * 10
        return rounded.quantize(Decimal('1.00'))

    def save_voucher_and_items(form, formset, voucher):
        with transaction.atomic():
            # Store original quantities for stock adjustment
            original_items = {}
            for item in voucher.items.all():
                original_items[item.id] = {
                    'quantity': item.quantity,
                    'product_id': item.product.id,
                    'phase': item.product.phase if hasattr(item.product, 'phase') else None
                }

            # Save the voucher form
            updated_voucher = form.save(commit=False)
            updated_voucher.save()

            total_subtotal = total_cgst = total_sgst = total_igst = grand_total = 0
            processed_items = set()

            for item_form in formset:
                if item_form.cleaned_data.get('DELETE', False):
                    # Handle deleted items - reverse stock changes
                    if item_form.instance.id:
                        original_item = original_items.get(item_form.instance.id)
                        if original_item:
                            product = Product.objects.get(id=original_item['product_id'])
                            if voucher_type == 'Seller_Voucher':
                                product.stock_quantity -= int(original_item['quantity'])
                            elif voucher_type == 'Buyer_Voucher':
                                product.stock_quantity += int(original_item['quantity'])
                            elif voucher_type == 'Job_Work_Voucher':
                                product.stock_quantity -= int(original_item['quantity'])
                                if original_item['phase']:
                                    product.phase = None
                            product.save()
                        item_form.instance.delete()
                    continue

                item = item_form.save(commit=False)
                item.voucher = updated_voucher

                # Calculate amounts
                item.subtotal = item.quantity * item.price_per_unit
                item.cgst_amount = (item.subtotal * item.cgst_percent) / 100
                item.sgst_amount = (item.subtotal * item.sgst_percent) / 100
                item.igst_amount = (item.subtotal * item.igst_percent) / 100
                item.total_amount = item.subtotal + item.cgst_amount + item.sgst_amount + item.igst_amount

                # Handle stock adjustments
                if item.id in original_items:
                    # Existing item - adjust stock based on difference
                    original_quantity = original_items[item.id]['quantity']
                    quantity_diff = int(item.quantity) - int(original_quantity)
                    
                    if voucher_type == 'Seller_Voucher':
                        item.product.stock_quantity += quantity_diff
                    elif voucher_type == 'Buyer_Voucher':
                        item.product.stock_quantity -= quantity_diff
                    elif voucher_type == 'Job_Work_Voucher':
                        item.product.stock_quantity += quantity_diff
                        selected_phase = item_form.cleaned_data.get('phase', '')
                        if selected_phase:
                            item.product.phase = selected_phase
                    
                    item.product.save()
                else:
                    # New item - full stock adjustment
                    if voucher_type == 'Seller_Voucher':
                        item.product.stock_quantity += int(item.quantity)
                    elif voucher_type == 'Buyer_Voucher':
                        item.product.stock_quantity -= int(item.quantity)
                    elif voucher_type == 'Job_Work_Voucher':
                        item.product.stock_quantity += int(item.quantity)
                        selected_phase = item_form.cleaned_data.get('phase', '')
                        if selected_phase:
                            item.product.phase = selected_phase
                    
                    item.product.save()

                item.save()
                processed_items.add(item.id)

                total_subtotal += item.subtotal
                total_cgst += item.cgst_amount
                total_sgst += item.sgst_amount
                total_igst += item.igst_amount
                grand_total += item.total_amount

            # Handle items that were removed from the formset
            for item_id, original_item in original_items.items():
                if item_id not in processed_items:
                    product = Product.objects.get(id=original_item['product_id'])
                    if voucher_type == 'Seller_Voucher':
                        product.stock_quantity -= int(original_item['quantity'])
                    elif voucher_type == 'Buyer_Voucher':
                        product.stock_quantity += int(original_item['quantity'])
                    elif voucher_type == 'Job_Work_Voucher':
                        product.stock_quantity -= int(original_item['quantity'])
                        if original_item['phase']:
                            product.phase = None
                    product.save()

            if updated_voucher.freight_applicable and updated_voucher.freight_charge:
                grand_total += updated_voucher.freight_charge

            # Round to nearest 10 and calculate round-off
            rounded_grand_total = round_to_nearest_10_with_00(grand_total)
            round_off = rounded_grand_total - grand_total

            # Save all calculated values
            updated_voucher.total_subtotal = total_subtotal
            updated_voucher.total_cgst = total_cgst
            updated_voucher.total_sgst = total_sgst
            updated_voucher.total_igst = total_igst
            updated_voucher.round_off_on_sales = round_off
            updated_voucher.grand_total = rounded_grand_total
            updated_voucher.save()

        return updated_voucher

    if request.method == 'POST':
        form = VoucherForm(request.POST or None, instance=voucher, voucher_type=voucher_type)
        formset = VoucherProductItemFormSetNoExtra(request.POST, instance=voucher)

        if form.is_valid() and formset.is_valid():
            if voucher_type == 'Job_Work_Voucher':
                invalid_products = []
                for item_form in formset:
                    if item_form.cleaned_data.get('DELETE', False):
                        continue
                    product = item_form.cleaned_data.get('product')
                    if product and product.type_of_supply != 'raw_material':
                        invalid_products.append(product.name)

                if invalid_products:
                    messages.error(request,
                        "For Job Work Voucher, only products with type_of_supply as 'raw_material' are allowed. "
                        "Invalid products: " + ", ".join(invalid_products)
                    )
                else:
                    updated_voucher = save_voucher_and_items(form, formset, voucher)
                    return redirect('voucher_detail', voucher_id=updated_voucher.id)

            elif voucher_type == 'Buyer_Voucher':
                insufficient_stock_products = []
                for item_form in formset:
                    if item_form.cleaned_data.get('DELETE', False):
                        continue
                    product = item_form.cleaned_data.get('product')
                    quantity = item_form.cleaned_data.get('quantity')
                    
                    # Calculate net change in quantity
                    if item_form.instance.id:
                        original_quantity = item_form.instance.quantity
                        quantity_diff = int(quantity) - int(original_quantity)
                    else:
                        quantity_diff = int(quantity)
                    
                    if product and quantity_diff > 0 and product.stock_quantity < quantity_diff:
                        insufficient_stock_products.append(
                            f"{product.name} (available: {product.stock_quantity}, requested change: {quantity_diff})"
                        )

                if insufficient_stock_products:
                    messages.error(request,
                        "Insufficient stock for the following products: " + ", ".join(insufficient_stock_products)
                    )
                else:
                    updated_voucher = save_voucher_and_items(form, formset, voucher)
                    return redirect('voucher_detail', voucher_id=updated_voucher.id)

            else:
                updated_voucher = save_voucher_and_items(form, formset, voucher)
                return redirect('voucher_detail', voucher_id=updated_voucher.id)

    else:
        form = VoucherForm(instance=voucher, voucher_type=voucher_type)
        formset = VoucherProductItemFormSetNoExtra(instance=voucher)

    products = Product.objects.select_related('warehouse').all()
    product_data = {
        str(product.id): {
            "warehouse_id": str(product.warehouse.id) if product.warehouse else ""
        }
        for product in products
    }

    return render(request, 'voucher_form_with_items.html', {
        'form': form,
        'formset': formset,
        'voucher_type': voucher_type,
        'product_data': json.dumps(product_data),
        'preview_voucher_number': voucher.voucher_number,
        'is_edit': True,
        'voucher': voucher
    })

def custom_404_view(request, exception):
    return render(request, '404.html', status=404)

def custom_500_view(request):
    return render(request, '500.html', status=500)


from .models import Account, Transaction
from .forms import AccountForm, TransactionForm


def account_list(request):
    accounts = Account.objects.all()
    return render(request, 'account_list.html', {'accounts': accounts})

def account_create(request):
    form = AccountForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('account_list')
    return render(request, 'account_form.html', {'form': form})

def account_update(request, pk):
    account = get_object_or_404(Account, pk=pk)
    form = AccountForm(request.POST or None, instance=account)
    if form.is_valid():
        form.save()
        return redirect('account_list')
    return render(request, 'account_form.html', {'form': form})

def account_delete(request, pk):
    account = get_object_or_404(Account, pk=pk)
    if request.method == 'POST':
        account.delete()
        return redirect('account_list')
    return render(request, 'account_confirm_delete.html', {'account': account})



def transaction_list(request):
    query = request.GET.get('q', '')
    txn_type = request.GET.get('type', '')

    transactions = Transaction.objects.all()

    if query:
        transactions = transactions.filter(
            Q(transaction_voucher_number__icontains=query) |
            Q(transaction_id__icontains=query) |
            Q(transaction_voucher_type__icontains=query) |
            Q(transaction_type__icontains=query) |
            Q(method_of_adjustment__icontains=query) |
            Q(account__account_holder_name__icontains=query) |
            Q(party__name__icontains=query) |
            Q(voucher__voucher_number__icontains=query) |
            Q(remarks__icontains=query)
        )

    if txn_type:
        transactions = transactions.filter(transaction_voucher_type=txn_type)

    transactions = transactions.order_by('-created_at')

    return render(request, 'transaction_list.html', {
        'transactions': transactions,
        'query': query,
    })



def select_transaction_type(request):
    return render(request, 'select_transaction_type.html')

from django.core.exceptions import ValidationError
from django.contrib import messages

def transaction_create(request, voucher_type):
    if request.method == 'POST':
        post_data = request.POST.copy()
        post_data['transaction_voucher_type'] = voucher_type
        form = TransactionForm(post_data, voucher_type=voucher_type)

        if form.is_valid():
            try:
                transaction = form.save()
                return redirect('transaction_detail', pk=transaction.pk)
            except ValidationError as e:
                # Add model clean() errors to form's non-field errors
                form.add_error(None, e.message)

    else:
        form = TransactionForm(voucher_type=voucher_type)

    return render(request, 'transaction_form.html', {
        'form': form,
        'voucher_type': voucher_type.capitalize()
    })

from django.http import JsonResponse
from django.views.decorators.http import require_GET

@require_GET
def get_vouchers(request):
    party_id = request.GET.get('party_id')
    if party_id:
        vouchers = Voucher.objects.filter(party_id=party_id).values('id', 'voucher_number')
        return JsonResponse(list(vouchers), safe=False)
    return JsonResponse([], safe=False)

from django.http import HttpResponse
from django.views import View

class LoadVouchersView(View):
    def get(self, request, *args, **kwargs):
        party_id = request.GET.get('party_id')
        vouchers = Voucher.objects.none()
        
        if party_id:
            vouchers = Voucher.objects.filter(party_id=party_id)
        
        options = ['<option value="">---------</option>']
        for voucher in vouchers:
            options.append(f'<option value="{voucher.id}">{voucher.voucher_number}</option>')
        return HttpResponse(''.join(options))
    
    
from django.shortcuts import render, get_object_or_404
from num2words import num2words
from .models import Transaction  # or adjust import as needed

def transaction_detail(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    
    amount_in_words = num2words(transaction.amount, to='currency', lang='en_IN').replace("euro", "rupees").replace("cents", "paise")

    context = {
        'transaction': transaction,
        'voucher_type_display': transaction.get_transaction_voucher_type_display(),
        'transaction_type_display': transaction.get_transaction_type_display(),
        'method_display': transaction.get_method_of_adjustment_display(),
        'amount_in_words': amount_in_words.title() + " Only",
    }
    return render(request, 'transaction_detail.html', context)

# views.py
from django.http import JsonResponse
from .models import Account

def get_account_balance(request):
    account_id = request.GET.get('account_id')
    try:
        account = Account.objects.get(id=account_id)
        return JsonResponse({'balance': float(account.available_balance)})
    except Account.DoesNotExist:
        return JsonResponse({'balance': 0})

from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from .models import Account, Transaction

def account_transaction_history(request, account_id):
    account = get_object_or_404(Account, id=account_id)

    transactions = Transaction.objects.filter(
        Q(account=account) | Q(recipient_account=account)
    ).order_by('-date', '-created_at')

    query = request.GET.get('search', '').lower()
    date_query = request.GET.get('date', '')

    if date_query:
        transactions = transactions.filter(date=date_query)

    history = []
    for txn in transactions:
        if txn.transaction_voucher_type == 'receipt' and txn.account == account:
            effect = 'Credit'
        elif txn.transaction_voucher_type == 'payment' and txn.account == account:
            effect = 'Debit'
        elif txn.transaction_voucher_type == 'contra':
            if txn.account == account:
                effect = 'Debit'
            elif txn.recipient_account == account:
                effect = 'Credit'
            else:
                effect = 'N/A'
        else:
            effect = 'N/A'

        record = {
            'date': txn.date,
            'voucher_number': txn.transaction_voucher_number,
            'type': txn.get_transaction_voucher_type_display(),
            'amount': txn.amount,
            'effect': effect,
            'party': txn.party.name if txn.party else '',
            'remarks': txn.remarks,
        }

        # Unified search filter
        combined_data = f"{record['date']} {record['voucher_number']} {record['type']} {record['effect']} {record['amount']} {record['party']}".lower()
        if query in combined_data:
            history.append(record)

    return render(request, 'account_transaction_history.html', {
        'account': account,
        'history': history,
        'search_query': query,
        'date_query': date_query,
    })


from django.forms import formset_factory
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import ValidationError
from django.shortcuts import render, get_object_or_404, redirect
from .models import Product, ProductComponent
from .forms import ProductComponentForm

@login_required
def add_finished_product_stock(request, pk):
    finished_product = get_object_or_404(Product, pk=pk)
    ProductComponentFormSet = formset_factory(ProductComponentForm, extra=1, can_delete=True)

    if finished_product.group.name.lower() != 'finished products':
        messages.error(request, "This is not a finished product.")
        return redirect('product_list')

    if request.method == 'POST':
        formset = ProductComponentFormSet(request.POST)
        try:
            quantity_to_add = int(request.POST.get('quantity_to_add', 0))
        except ValueError:
            quantity_to_add = 0

        if formset.is_valid() and quantity_to_add > 0:
            try:
                with transaction.atomic():
                    for form in formset:
                        if form.cleaned_data.get('DELETE'):
                            continue

                        component = form.cleaned_data.get('component_product')
                        qty_used = form.cleaned_data.get('quantity_used')

                        if not component or not qty_used:
                            continue

                        used_qty = qty_used * quantity_to_add

                        if component.stock_quantity < used_qty:
                            raise ValidationError(f"Insufficient stock for {component.name}")

                        component.stock_quantity -= used_qty
                        component.save()

                        ProductComponent.objects.create(
                            finished_product=finished_product,
                            component_product=component,
                            quantity_used=qty_used
                        )

                    finished_product.stock_quantity += quantity_to_add
                    finished_product.save()

                    messages.success(request, f"{quantity_to_add} units of {finished_product.name} added.")
                    return redirect('product_list')

            except ValidationError as e:
                messages.error(request, str(e))

    else:
        formset = ProductComponentFormSet()

    return render(request, 'add_finished_product_stock.html', {
        'finished_product': finished_product,
        'formset': formset,
    })


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Voucher
from decimal import Decimal
import calendar
from datetime import datetime

def export_sales_summary_excel(request, year, month):
    month_name = calendar.month_name[int(month)]
    vouchers = Voucher.objects.filter(
        voucher_type='Buyer_Voucher',
        created_at__year=year,
        created_at__month=month
    ).order_by('created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = f"Sales {month_name} {year}"

    # === Define styles ===
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True)
    currency_format = '₹#,##0.00'
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # === Company Header Section ===
    company_info = [
        "VVM AGRO INDUSTRIES (2020-26)",
        "SURVEY NO. 247/AA AND 249/A1, KONDARPUR ROAD",
        "BESIDE TSIIC KALLAKAL, MUPPYREDDY PALLY VILLAGE",
        "MANOHARABAD MANDAL, MEDAK DIST.",
        "Contact : 9246565834",
        "Sales Register",
        f"1-{month_name[:3]}-{year} to {calendar.monthrange(int(year), int(month))[1]}-{month_name[:3]}-{year}"
    ]

    for line in company_info:
        row_idx = ws.max_row + 1
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.font = title_font
        cell.alignment = center_align

    ws.append([])  # Empty row

    # === Table Header ===
    headers = [
        "Date", "Particulars", "Buyer", "Buyer Address", "Voucher Type", "Voucher No.",
        "GSTIN/UIN", "Value", "Gross Total", "INTER STATE SALES", "Output IGST",
        "SALES", "Output CGST", "Output SGST", "ROUND OFF ON SALES", "FREIGHT"
    ]
    ws.append(headers)
    header_row = ws.max_row

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    ws.append([])  # Empty row

    # === Data Rows ===
    for voucher in vouchers:
        party = voucher.party
        cgst = voucher.total_cgst or Decimal('0.00')
        sgst = voucher.total_sgst or Decimal('0.00')
        igst = voucher.total_igst or Decimal('0.00')
        round_off = voucher.round_off_on_sales or Decimal('0.00')
        freight = voucher.freight_charge or Decimal('0.00')
        subtotal = voucher.total_subtotal or Decimal('0.00')
        total = voucher.grand_total or Decimal('0.00')
        is_inter_state = bool(igst > 0)
        interstate_sales = subtotal if is_inter_state else ''
        local_sales = subtotal if not is_inter_state else ''

        row_data = [
            voucher.created_at.strftime('%d-%m-%Y'),
            party.name,
            party.name,
            party.address,
            "Sales",
            voucher.voucher_number,
            party.gstin_uin_number,
            subtotal,
            total,
            interstate_sales,
            igst,
            local_sales,
            cgst,
            sgst,
            round_off,
            freight
        ]

        ws.append(row_data)
        current_row = ws.max_row

        # Apply formatting to each cell
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_index)
            cell.border = thin_border
            if isinstance(value, Decimal) or isinstance(value, float):
                cell.number_format = currency_format
                cell.alignment = right_align = Alignment(horizontal='right')
            else:
                cell.alignment = left_align

    # === Auto-adjust column widths ===
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(14, max_length + 2)

    # === Freeze header row (after company section) ===
    ws.freeze_panes = f"A{header_row + 2}"

    # === Return file response ===
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Sales_Summary_{month_name}_{year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


import datetime as dt
import calendar
from django.shortcuts import render

def download_sales_summary_page(request):
    current_year = dt.datetime.now().year
    years = [current_year - i for i in range(100)]  # Last 5 years
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    return render(request, 'export_sales_summary.html', {'years': years, 'months': months})


from django.shortcuts import render
import datetime

def hsn_summary_form(request):
    current_year = datetime.datetime.now().year
    years = list(range(current_year, current_year - 100, -1))  # Last 10 years
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'),
        (4, 'April'), (5, 'May'), (6, 'June'),
        (7, 'July'), (8, 'August'), (9, 'September'),
        (10, 'October'), (11, 'November'), (12, 'December')
    ]
    return render(request, 'export_hsn_summary.html', {'years': years, 'months': months})

def export_hsn_gst_summary_excel(request, year, month):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    import calendar
    from .models import VoucherProductItem

    month_name = calendar.month_name[int(month)]
    start_date = f"1-{month_name[:3]}-{year}"
    end_day = calendar.monthrange(int(year), int(month))[1]
    end_date = f"{end_day}-{month_name[:3]}-{year}"

    items = VoucherProductItem.objects.filter(
        voucher__voucher_type='Buyer_Voucher',
        voucher__created_at__year=year,
        voucher__created_at__month=month
    ).select_related('product')

    summary = defaultdict(lambda: {
        'description': '',
        'unit': '',
        'quantity': 0,
        'taxable': 0,
        'igst': 0,
        'cgst': 0,
        'sgst': 0,
        'cess': 'NA',
        'rate': 0
    })

    for item in items:
        key = item.product.hsn_sac
        summary[key]['description'] = item.product.name
        summary[key]['unit'] = item.product.unit_of_measurement
        summary[key]['quantity'] += float(item.quantity)
        summary[key]['taxable'] += float(item.subtotal)
        summary[key]['igst'] += float(item.igst_amount)
        summary[key]['cgst'] += float(item.cgst_amount)
        summary[key]['sgst'] += float(item.sgst_amount)
        summary[key]['rate'] = float(item.igst_percent + item.cgst_percent + item.sgst_percent)

    wb = Workbook()
    ws = wb.active
    ws.title = f"HSN Summary {month_name} {year}"

    # Styles
    title_font = Font(size=14, bold=True)
    header_font = Font(size=12, bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center')
    left = Alignment(horizontal='left')
    right = Alignment(horizontal='right')
    currency_format = '₹#,##0.00'
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Company Info
    top_lines = [
        "VVM AGRO INDUSTRIES (2020-26)",
        "SURVEY NO. 247/AA AND 249/A1, KONDARPUR ROAD",
        "BESIDE TSIIC KALLAKAL, MUPPYREDDY PALLY VILLAGE",
        "MANOHARABAD MANDAL, MEDAK DIST.",
        "Contact : 9246565834",
        "GSTR-1 Reconciliation - HSN/SAC Summary",
        f"{start_date} to {end_date}",
        "GST Registration:",
        "36AANFV9684N1ZW (Comparison of Books & Portal Values)",
        "HSN/SAC Summary - HSN/SAC"
    ]

    for text in top_lines:
        row = ws.max_row + 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = title_font
        cell.alignment = center

    ws.append([])

    # Header Row
    headers = [
        "HSN/SAC", "Description", "UQC", "Quantity",
        "Total Amount (₹)", "Tax Rate (%)", "Taxable Amount (₹)",
        "IGST (₹)", "CGST (₹)", "SGST/UTGST (₹)", "Cess", "Tax Amount (₹)"
    ]
    ws.append(headers)
    r = ws.max_row
    for c in range(1, 13):
        cell = ws.cell(row=r, column=c)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    # Data Rows
    for hsn, data in summary.items():
        total_tax = data['igst'] + data['cgst'] + data['sgst']
        row = [
            hsn,
            data['description'],
            data['unit'],
            round(data['quantity'], 2),
            data['taxable'] + total_tax,
            data['rate'],
            data['taxable'],
            data['igst'],
            data['cgst'],
            data['sgst'],
            data['cess'],  # 'NA'
            total_tax
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.border = border

            # Format numeric ₹ columns
            if c in [5, 7, 8, 9, 10, 12] and isinstance(cell.value, (int, float)):
                cell.alignment = right
                cell.number_format = currency_format
            elif c == 6:  # Tax Rate (%)
                cell.alignment = center
                cell.number_format = '0.00'
            else:
                cell.alignment = left

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(15, max_len + 2)

    # Output response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"HSN_GST_Summary_{month_name}_{year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import Product
import pytz

def export_products_excel(request):
    # Create a new workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Product List"

    # Styling
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')

    # Header row
    headers = [
        "Name", "Group", "Price per Unit (₹)", "Unit of Measurement", "Stock Quantity",
        "Warehouse", "Stock Limit", "HSN/SAC", "HSN/SAC Details", "Source of HSN/SAC",
        "HSN/SAC Description", "Type of Supply", "Phase", "Description", "Created At (IST)"
    ]
    ws.append(headers)

    for col_num, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align

    # Data rows
    for product in Product.objects.select_related('group', 'warehouse'):
        row = [
            product.name,
            product.group.name if product.group else '',
            float(product.price_per_unit),
            product.unit_of_measurement,
            product.stock_quantity,
            product.warehouse.name if product.warehouse else '',
            product.stock_limit,
            product.hsn_sac,
            product.hsn_sac_details,
            product.hsn_sac_source_of_details,
            product.hsn_sac_description,
            product.get_type_of_supply_display(),
            product.get_phase_display() if product.phase else '',
            product.description,
            product.created_at.astimezone(pytz.timezone('Asia/Kolkata')).strftime('%Y-%m-%d %H:%M:%S'),
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = left_align

    # Auto column width
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(15, max_length + 2)

    # Prepare HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "Product_List.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Party

def export_party_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Party List"

    # Styling
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')

    # Header Row
    headers = [
        "Party Group", "Name", "GSTIN/UIN", "Address", "Location",
        "Pincode", "State", "Phone", "Email", "Registration Type"
    ]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Data Rows
    for party in Party.objects.select_related('group'):
        row = [
            party.group.name if party.group else '',
            party.name,
            party.gstin_uin_number,
            party.address,
            party.location,
            party.pincode,
            party.state,
            party.phone,
            party.email,
            party.registration_type  # <-- Added here
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = left_align

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(15, max_length + 2)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "Party_List.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



from django.shortcuts import render, get_object_or_404, redirect
from .forms import TransactionForm
from .models import Transaction

def transaction_edit(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    voucher_type = transaction.transaction_voucher_type

    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction, voucher_type=voucher_type)
        if form.is_valid():
            form.save()
            return redirect('transaction_detail', pk=transaction.pk)
    else:
        form = TransactionForm(instance=transaction, voucher_type=voucher_type)

    return render(request, 'transaction_form.html', {
        'form': form,
        'voucher_type': voucher_type.capitalize()
    })
